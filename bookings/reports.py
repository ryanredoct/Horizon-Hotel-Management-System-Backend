from io import BytesIO

import openpyxl
from django.http import JsonResponse, HttpResponse
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from hotels.models import RoomCategory
from .models import Room, Booking  # Assuming Room model exists


def total_rooms_report(request):
    try:
        # Get all the rooms data
        rooms = Room.objects.all()

        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Total Rooms Report"

        # Add headers with styling
        headers = ["Room Number", "Room Category", "Available"]
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        report_title = "Rooms Report"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=report_title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        header_row = 2
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Start writing data below the headers
        row_num = header_row + 1
        for data in rooms:
            ws.cell(row=row_num, column=1, value=data.room_number)  # Example: Room number
            ws.cell(row=row_num, column=2, value=data.category.name)  # Example: Room type
            ws.cell(row=row_num, column=3, value=data.is_available)  # Example: Status
            # ws.cell(row=row_num, column=4, value=data.price)  # Example: Price
            row_num += 1

        # Adjust column widths
        for col_num, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 18

        # Save workbook to a BytesIO stream
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Prepare response with Excel file
        response = HttpResponse(output,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="total_rooms_report.xlsx"'
        return response

    except Exception as e:
        return JsonResponse({'success': False, 'error_message': f'Error generating total rooms report Excel: {str(e)}'})


def available_rooms_report(request):
    try:
        # Get all the rooms data
        rooms = Room.objects.filter(is_available=True)

        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Available Rooms Report"

        # Add headers with styling
        headers = ["Room Number", "Room Category", "Available"]
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        report_title = "Available Rooms Report"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=report_title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        header_row = 2
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Start writing data below the headers
        row_num = header_row + 1
        for data in rooms:
            ws.cell(row=row_num, column=1, value=data.room_number)  # Example: Room number
            ws.cell(row=row_num, column=2, value=data.category.name)  # Example: Room type
            ws.cell(row=row_num, column=3, value=data.is_available)  # Example: Status
            # ws.cell(row=row_num, column=4, value=data.price)  # Example: Price
            row_num += 1

        # Adjust column widths
        for col_num, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 18

        # Save workbook to a BytesIO stream
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Prepare response with Excel file
        response = HttpResponse(output,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="available_rooms_report.xlsx"'
        return response

    except Exception as e:
        return JsonResponse(
            {'success': False, 'error_message': f'Error generating available rooms report Excel: {str(e)}'})


def room_categories_report(request):
    try:
        # Get all the rooms data
        room_categories = RoomCategory.objects.all()

        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Room Categories Report"

        # Add headers with styling
        headers = ["Name", "Capacity", "Price per Night"]
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        report_title = "Room Categories Report"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=report_title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        header_row = 2
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Start writing data below the headers
        row_num = header_row + 1
        for data in room_categories:
            ws.cell(row=row_num, column=1, value=data.name)  # Example: Room number
            ws.cell(row=row_num, column=2, value=data.capacity)  # Example: Room type
            ws.cell(row=row_num, column=3, value=data.price_per_night)  # Example: Status
            # ws.cell(row=row_num, column=4, value=data.price)  # Example: Price
            row_num += 1

        # Adjust column widths
        for col_num, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 18

        # Save workbook to a BytesIO stream
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Prepare response with Excel file
        response = HttpResponse(output,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="room_categories_report.xlsx"'
        return response

    except Exception as e:
        return JsonResponse(
            {'success': False, 'error_message': f'Error generating room categories report Excel: {str(e)}'})


def bookings_report(request):
    try:
        # Get all the rooms data
        bookings = Booking.objects.all()

        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bookings Report"

        # Add headers with styling
        headers = ["Booking No.", "Customer", "Customer ID", "Check In", "Check Out", "Status", "Total"]
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        report_title = "Bookings Report"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=report_title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        header_row = 2
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Start writing data below the headers
        row_num = header_row + 1
        for data in bookings:
            ws.cell(row=row_num, column=1, value=data.booking_number)  # Example: Room number
            ws.cell(row=row_num, column=2, value=data.customer_name)  # Example: Room type
            ws.cell(row=row_num, column=3, value=data.customer_id_number)  # Example: Status
            ws.cell(row=row_num, column=4, value=data.check_in)  # Example: Price
            ws.cell(row=row_num, column=5, value=data.check_out)  # Example: Price
            ws.cell(row=row_num, column=6, value=data.status)  # Example: Price
            ws.cell(row=row_num, column=7, value=data.total_price)  # Example: Price
            row_num += 1

        # Adjust column widths
        for col_num, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 18

        # Save workbook to a BytesIO stream
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Prepare response with Excel file
        response = HttpResponse(output,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="bookings_report.xlsx"'
        return response

    except Exception as e:
        return JsonResponse(
            {'success': False, 'error_message': f'Error generating room categories report Excel: {str(e)}'})


def revenue_report(request):
    try:
        # Get all the rooms data
        bookings = Booking.objects.filter(status='CheckOut')

        # Calculate total revenue
        total_revenue = sum(booking.total_price for booking in bookings)

        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Revenue Report"

        # Add headers with styling
        headers = ["Booking No.", "Customer", "Customer ID", "Check In", "Check Out", "Status", "Total"]
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        report_title = "Revenue Report"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=report_title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        header_row = 2
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Start writing data below the headers
        row_num = header_row + 1
        for data in bookings:
            ws.cell(row=row_num, column=1, value=data.booking_number)
            ws.cell(row=row_num, column=2, value=data.customer_name)
            ws.cell(row=row_num, column=3, value=data.customer_id_number)
            ws.cell(row=row_num, column=4, value=data.check_in)
            ws.cell(row=row_num, column=5, value=data.check_out)
            ws.cell(row=row_num, column=6, value=data.status)
            ws.cell(row=row_num, column=7, value=data.total_price)
            row_num += 1

        # Add Total Revenue Row
        total_label_cell = ws.cell(row=row_num, column=1, value="Total Revenue")
        total_label_cell.font = Font(bold=True)
        total_label_cell.alignment = Alignment(horizontal="right")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)

        total_value_cell = ws.cell(row=row_num, column=7, value=total_revenue)
        total_value_cell.font = Font(bold=True)
        total_value_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        # Adjust column widths
        for col_num, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 18

        # Save workbook to a BytesIO stream
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Prepare response with Excel file
        response = HttpResponse(output,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="revenue_report.xlsx"'
        return response

    except Exception as e:
        return JsonResponse({'success': False, 'error_message': f'Error generating revenue report Excel: {str(e)}'})
